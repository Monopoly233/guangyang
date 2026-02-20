from fastapi import APIRouter, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from typing import List, Optional
import pandas as pd
import io
import logging
import traceback
import os
import asyncio
from urllib.parse import quote

from starlette.concurrency import run_in_threadpool

from .utils import (
    guess_primary_key_column,
    prepare_table_data,
    compare_excel_tables,
    compare_excel_tables_df,
    read_excel_bytes,
    _normalize_scalar_for_compare,
)
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)

router = APIRouter(prefix="", tags=["compare"])

_MAX_CONCURRENCY = int(os.getenv("COMPARE_MAX_CONCURRENCY", "2") or "2")
_COMPARE_SEM = asyncio.Semaphore(max(1, _MAX_CONCURRENCY))

def _compare_json_impl(content1: bytes, content2: bytes, file1name: str, file2name: str):
    try:
        df1 = read_excel_bytes(content1, file1name or "")
        df2 = read_excel_bytes(content2, file2name or "")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"读取Excel失败: {str(e)}")

    onlyid = guess_primary_key_column(df1)
    if not onlyid:
        raise HTTPException(status_code=400, detail='无法猜测主键列，请确保包含明显的编号列')

    if onlyid not in df1.columns or onlyid not in df2.columns:
        raise HTTPException(status_code=400, detail=f'Excel文件中必须同时包含"{onlyid}"列')

    try:
        _ = set(df1[onlyid].astype(str).dropna())
        _ = set(df2[onlyid].astype(str).dropna())
    except Exception as e:
        raise HTTPException(status_code=400, detail=f'处理序列号时出错: {str(e)}')

    reduced_tbl, increased_tbl, different_tbl = compare_excel_tables(
        df1, df2, onlyid, file1name, file2name
    )

    response_data = {
        'reduced': reduced_tbl,
        'increased': increased_tbl,
        'different': different_tbl,
    }

    if not any(response_data.values()):
        return {'error': '没有找到任何差异数据'}
    return response_data


def _compare_export_bytes_impl(content1: bytes, content2: bytes, file1name: str, file2name: str) -> bytes:
    try:
        df1 = read_excel_bytes(content1, file1name or "")
        df2 = read_excel_bytes(content2, file2name or "")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"读取Excel失败: {str(e)}")

    onlyid = guess_primary_key_column(df1)
    if not onlyid:
        raise HTTPException(status_code=400, detail='无法猜测主键列，请确保包含明显的编号列')
    if onlyid not in df1.columns or onlyid not in df2.columns:
        raise HTTPException(status_code=400, detail=f'Excel文件中必须同时包含"{onlyid}"列')

    reduced_df, increased_df, _different_df = compare_excel_tables_df(
        df1=df1,
        df2=df2,
        key=onlyid,
        file1name=file1name,
        file2name=file2name,
    )

    # 将结果写入 xlsx：固定三张表（增加项/减少项/变动项目）
    output = io.BytesIO()
    wb = Workbook(write_only=True)
    base1 = _sheet_base_name(file1name)
    base2 = _sheet_base_name(file2name)
    used_names = set()
    # 方向：文件2 相比 文件1 的增加/减少
    ws_inc = wb.create_sheet(_unique_sheet_name(f"{base2}相比{base1}增加", used_names))
    ws_red = wb.create_sheet(_unique_sheet_name(f"{base2}相比{base1}减少", used_names))
    ws_diff = wb.create_sheet(_unique_sheet_name("变动项目", used_names))

    def write_df(ws, df: Optional[pd.DataFrame], empty_msg: str):
        if df is None or df.empty:
            ws.append([empty_msg])
            return
        # openpyxl 不能直接写入 list/dict 等复杂对象（例如 “不同列” 是 list）
        dfw = df.copy()
        for c in dfw.columns:
            dfw[c] = dfw[c].map(
                lambda x: ",".join(map(str, x)) if isinstance(x, (list, tuple, set)) else x
            )

        for r in dataframe_to_rows(dfw, index=False, header=True):
            ws.append(r)

    write_df(ws_inc, increased_df, "无增加项")
    write_df(ws_red, reduced_df, "无减少项")
    _write_diff_side_by_side(ws_diff, df1=df1, df2=df2, key=onlyid, file1name=file1name, file2name=file2name)

    wb.save(output)
    return output.getvalue()

def _sheet_base_name(filename: str) -> str:
    name = (filename or "").strip()
    if not name:
        return "文件"
    name = os.path.basename(name)
    if "." in name:
        name = name.rsplit(".", 1)[0]
    name = name.strip()
    return name or "文件"


def _safe_sheet_name(name: str) -> str:
    """
    Excel 工作表名限制：
    - 不能包含 : \\ / ? * [ ]
    - 最大长度 31
    """
    s = (name or "").strip() or "Sheet"
    for ch in [":", "\\", "/", "?", "*", "[", "]"]:
        s = s.replace(ch, "_")
    if len(s) > 31:
        s = s[:31]
    return s


def _unique_sheet_name(name: str, used: set) -> str:
    base = _safe_sheet_name(name)
    cand = base
    i = 2
    while cand in used:
        suffix = f"_{i}"
        max_len = 31 - len(suffix)
        cand = (base[:max_len] if len(base) > max_len else base) + suffix
        i += 1
    used.add(cand)
    return cand


def _safe_cell_value(v):
    try:
        if pd.isna(v):
            return ""
    except Exception:
        pass
    if isinstance(v, (list, tuple, set)):
        return ",".join(map(str, v))
    return v


def _ordered_union_cols(df1: pd.DataFrame, df2: pd.DataFrame, key: str):
    cols1 = list(df1.columns)
    cols2 = list(df2.columns)
    out = [c for c in cols1 if c != key]
    for c in cols2:
        if c == key:
            continue
        if c not in out:
            out.append(c)
    return out


def _write_diff_side_by_side(ws, df1: pd.DataFrame, df2: pd.DataFrame, key: str, file1name: str, file2name: str):
    """
    变动项目：同主键两侧并列在同一行，并将不同的 cell 标红。
    列顺序：按 file1 原始顺序，再补 file2 新列（顺序保持不变）。
    """
    if df1 is None or df2 is None or df1.empty or df2.empty:
        ws.append(["无变动项目"])
        return

    # normalize key for stable alignment
    df1_c = df1.copy()
    df2_c = df2.copy()
    df1_c[key] = df1_c[key].map(_normalize_scalar_for_compare)
    df2_c[key] = df2_c[key].map(_normalize_scalar_for_compare)
    df1_c = df1_c[df1_c[key].astype(str).str.strip() != ""]
    df2_c = df2_c[df2_c[key].astype(str).str.strip() != ""]
    df1_c = df1_c.drop_duplicates(subset=[key], keep="first").set_index(key)
    df2_c = df2_c.drop_duplicates(subset=[key], keep="first").set_index(key)

    common_idx = df1_c.index.intersection(df2_c.index)
    if len(common_idx) == 0:
        ws.append(["无变动项目"])
        return

    ordered_cols = _ordered_union_cols(df1, df2, key=key)
    a = df1_c.loc[common_idx].reindex(columns=ordered_cols)
    b = df2_c.loc[common_idx].reindex(columns=ordered_cols)

    # fast diff to narrow down candidates
    try:
        diff_fast = a.ne(b)
        both_na = a.isna() & b.isna()
        diff_fast = diff_fast & (~both_na)
        cand_rows = diff_fast.any(axis=1)
        cand_cols = diff_fast.loc[cand_rows].any(axis=0) if cand_rows.any() else None
    except Exception:
        cand_rows = pd.Series(True, index=a.index)
        cand_cols = None

    if not cand_rows.any():
        ws.append(["无变动项目"])
        return

    a_sub = a.loc[cand_rows]
    b_sub = b.loc[cand_rows]

    cols_to_check = ordered_cols
    if cand_cols is not None and getattr(cand_cols, "any", lambda: False)():
        _cols = [c for c, v in cand_cols.items() if bool(v)]
        if _cols:
            cols_to_check = _cols

    a_cmp = a_sub[cols_to_check].apply(lambda s: s.map(_normalize_scalar_for_compare))
    b_cmp = b_sub[cols_to_check].apply(lambda s: s.map(_normalize_scalar_for_compare))
    diff_mask = a_cmp.ne(b_cmp)
    rows_with_diff = diff_mask.any(axis=1)
    if not rows_with_diff.any():
        ws.append(["无变动项目"])
        return

    # header
    header = [key]
    for c in ordered_cols:
        header.append(f"{c}（{file1name or '文件1'}）")
        header.append(f"{c}（{file2name or '文件2'}）")
    ws.append(header)

    red_fill = PatternFill("solid", fgColor="FFFFC7CE")  # light red
    red_font = Font(color="FF9C0006")

    keys = list(diff_mask.index[rows_with_diff])
    for k in keys:
        row = []
        row.append(WriteOnlyCell(ws, value=_safe_cell_value(k)))
        for c in ordered_cols:
            va = _safe_cell_value(a_sub.at[k, c]) if c in a_sub.columns else ""
            vb = _safe_cell_value(b_sub.at[k, c]) if c in b_sub.columns else ""
            is_diff = False
            if c in diff_mask.columns:
                try:
                    is_diff = bool(diff_mask.at[k, c])
                except Exception:
                    is_diff = False

            ca = WriteOnlyCell(ws, value=va)
            cb = WriteOnlyCell(ws, value=vb)
            if is_diff:
                ca.fill = red_fill
                cb.fill = red_fill
                ca.font = red_font
                cb.font = red_font
            row.append(ca)
            row.append(cb)
        ws.append(row)


@router.post("/compare/")
async def compare_excel(file1: UploadFile = File(...), file2: UploadFile = File(...)):
    try:
        if file1 is None or file2 is None:
            raise HTTPException(status_code=400, detail='请上传两个Excel文件')

        content1 = await file1.read()
        content2 = await file2.read()

        async with _COMPARE_SEM:
            return await run_in_threadpool(_compare_json_impl, content1, content2, file1.filename or "", file2.filename or "")
    except HTTPException:
        raise
    except Exception:
        logger.error(f"请求处理出错: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail='服务器内部错误')

@router.post("/compare/export")
async def compare_excel_export(file1: UploadFile = File(...), file2: UploadFile = File(...)):
    try:
        if file1 is None or file2 is None:
            raise HTTPException(status_code=400, detail='请上传两个Excel文件')

        content1 = await file1.read()
        content2 = await file2.read()

        async with _COMPARE_SEM:
            out_bytes = await run_in_threadpool(
                _compare_export_bytes_impl,
                content1,
                content2,
                file1.filename or "",
                file2.filename or "",
            )
        output = io.BytesIO(out_bytes)
        filename = f"对比结果_{file1.filename}_vs_{file2.filename}.xlsx"
        encoded_name = quote(filename)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                # 同时提供 ASCII 回退与 RFC 5987 编码，避免非 ASCII 触发 latin-1 编码报错
                "Content-Disposition": f"attachment; filename=\"comparison_result.xlsx\"; filename*=UTF-8''{encoded_name}",
            },
        )
    except HTTPException:
        raise
    except Exception:
        logger.error(f"导出对比Excel出错: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail='服务器内部错误')